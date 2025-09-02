import csv
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QStackedWidget,
    QFileDialog, QInputDialog, QMessageBox, QDialog
)
from PySide6.QtCore import QDate
from openpyxl import Workbook

# Import your app modules
from bars_and_tabs.sidebar import Sidebar
from bars_and_tabs.dashboard import DashboardPage
from bars_and_tabs.repair_orders import RepairOrdersPage
from bars_and_tabs.reports import ReportsPage
from bars_and_tabs.settings import SettingsPage
from database import get_connection


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Tech Efficiency Tracker")
        self.resize(1100, 700)

        # --- Central widget & layouts ---
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QHBoxLayout(central_widget)

        # Main stacked area
        self.stack = QStackedWidget()
        layout.addWidget(self.stack, stretch=4)

        # --- Pages for the stack ---
        self.stack.addWidget(DashboardPage())      # index 0
        self.stack.addWidget(RepairOrdersPage())   # index 1
        self.stack.addWidget(ReportsPage())        # index 2
        self.stack.addWidget(SettingsPage())       # index 3

        # Sidebar (pass export callback)
        sidebar = Sidebar(self.stack, self.handle_export)
        layout.insertWidget(0, sidebar, stretch=1)

    # ---------------- Export Handler ----------------
    def handle_export(self):
        idx = self.stack.currentIndex()
        if idx == 1:   # Repair Orders
            self.export_repair_orders()
        elif idx == 2: # Reports
            self.export_reports()
        else:
            QMessageBox.information(self, "Export", "Export only works on ROs or Reports pages.")

    # ---------------- Repair Orders Export ----------------
    def export_repair_orders(self):
        choice, ok = QInputDialog.getItem(
            self, "Export ROs", "Choose export type:",
            ["All", "Open", "Closed", "Custom Date Range"], 0, False
        )
        if not ok:
            return

        query = "SELECT id, date, ro_number, estimator_id, stage, status, hours_total, hours_body, hours_refinish, hours_mechanical FROM repair_orders"
        params = []
        if choice == "Open":
            query += " WHERE status='Open'"
        elif choice == "Closed":
            query += " WHERE status='Closed'"
        elif choice == "Custom Date Range":
            from_date = QDate.currentDate().addDays(-90)
            to_date = QDate.currentDate()
            query += " WHERE date BETWEEN ? AND ?"
            params = [from_date.toString("yyyy-MM-dd"), to_date.toString("yyyy-MM-dd")]

        with get_connection() as conn:
            cur = conn.cursor()
            cur.execute(query, params)
            rows = cur.fetchall()

        # --- Save ROs CSV with Tech/Painter included ---
        file, _ = QFileDialog.getSaveFileName(self, "Save ROs CSV", "", "CSV Files (*.csv)")
        if not file:
            return

        with open(file, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Date", "RO#", "Estimator", "Tech", "Painter",
                "Stage", "Status", "Total Hours", "Body Hours", "Refinish Hours", "Mechanical Hours"
            ])
            for ro_id, date, ro_number, estimator_id, stage, status, hours_total, hours_body, hours_refinish, hours_mechanical in rows:
                qdate = QDate.fromString(date, "yyyy-MM-dd")
                date_fmt = qdate.toString("MM/dd/yyyy") if qdate.isValid() else date

                # Estimator
                cur.execute("SELECT full_name, nickname FROM employees WHERE id=?", (estimator_id,))
                est = cur.fetchone()
                estimator = (est[1] or est[0]) if est else "Unknown"

                # --- Look up Tech & Painter allocations ---
                cur.execute("""
                            SELECT e.full_name, e.nickname, a.role
                            FROM ro_hours_allocation a
                                     JOIN employees e ON a.employee_id = e.id
                            WHERE a.ro_id = ?
                            """, (ro_id,))
                allocs = cur.fetchall()
                techs, painters = [], []
                for fname, nname, role in allocs:
                    display = nname or fname
                    if role == "Tech":
                        techs.append(display)
                    elif role == "Painter":
                        painters.append(display)

                tech_str = ", ".join(techs) if techs else "Unassigned"
                painter_str = ", ".join(painters) if painters else "Unassigned"

                writer.writerow([
                    date_fmt, ro_number, estimator, tech_str, painter_str,
                    stage, status, hours_total, hours_body, hours_refinish, hours_mechanical
                ])

        QMessageBox.information(self, "Export", f"ROs exported to {file}")

    # ---------------- Reports Export ----------------
    def export_reports(self):
        choice, ok = QInputDialog.getItem(
            self, "Export Reports", "Choose export type:",
            ["All", "Last 90 Days", "Custom Date Range"], 1, False
        )
        if not ok:
            return

        where_h = ""
        where_c = ""
        params_h = []
        params_c = []

        if choice == "Last 90 Days":
            from_date = QDate.currentDate().addDays(-90).toString("yyyy-MM-dd")
            to_date = QDate.currentDate().toString("yyyy-MM-dd")
            where_h = "WHERE substr(h.date, 1, 10) BETWEEN ? AND ?"
            where_c = "WHERE substr(c.date, 1, 10) BETWEEN ? AND ?"
            params_h = [from_date, to_date]
            params_c = [from_date, to_date]

        elif choice == "Custom Date Range":
            from utilities.date_dialog import DatePickerDialog
            dlg1 = DatePickerDialog("Select Start Date", QDate.currentDate().addDays(-30), self)
            if dlg1.exec() != QDialog.Accepted: return
            from_date = dlg1.selected_date
            dlg2 = DatePickerDialog("Select End Date", QDate.currentDate(), self)
            if dlg2.exec() != QDialog.Accepted: return
            to_date = dlg2.selected_date
            where_h = "WHERE substr(h.date, 1, 10) BETWEEN ? AND ?"
            where_c = "WHERE substr(c.date, 1, 10) BETWEEN ? AND ?"
            params_h = [from_date.toString("yyyy-MM-dd"), to_date.toString("yyyy-MM-dd")]
            params_c = [from_date.toString("yyyy-MM-dd"), to_date.toString("yyyy-MM-dd")]

        file, _ = QFileDialog.getSaveFileName(self, "Save Reports", "", "Excel Files (*.xlsx)")
        if not file:
            return

        wb = Workbook()

        # --- Credited Hours ---
        ws1 = wb.active
        ws1.title = "Credited Hours"
        with get_connection() as conn:
            cur = conn.cursor()
            query = f"""
                SELECT e.full_name, e.nickname, SUM(c.hours)
                FROM credit_audit c
                JOIN employees e ON c.employee_id = e.id
                {where_c}
                GROUP BY e.id
            """
            cur.execute(query, params_c) if params_c else cur.execute(query)
            rows = cur.fetchall()
        ws1.append(["Employee", "Credited Hours"])
        for full_name, nickname, hrs in rows:
            display = nickname or full_name
            ws1.append([display, f"{hrs:.2f}"])

        # --- Efficiency ---
        ws2 = wb.create_sheet("Efficiency")
        with get_connection() as conn:
            cur = conn.cursor()
            query1 = f"""
                SELECT e.full_name, e.nickname, SUM(h.hours_worked)
                FROM employee_hours h
                JOIN employees e ON h.employee_id = e.id
                {where_h}
                GROUP BY e.id
            """
            cur.execute(query1, params_h) if params_h else cur.execute(query1)
            worked_map = {fn: (nn, total or 0) for fn, nn, total in cur.fetchall()}

            query2 = f"""
                SELECT e.full_name, e.nickname, SUM(c.hours)
                FROM credit_audit c
                JOIN employees e ON c.employee_id = e.id
                {where_c}
                GROUP BY e.id
            """
            cur.execute(query2, params_c) if params_c else cur.execute(query2)
            credits_map = {fn: (nn, total or 0) for fn, nn, total in cur.fetchall()}

        employees = set(worked_map.keys()) | set(credits_map.keys())
        ws2.append(["Employee", "Hours Worked", "Credited Hours", "Efficiency"])
        for full_name in sorted(employees):
            nn_w, worked = worked_map.get(full_name, (None, 0))
            nn_c, produced = credits_map.get(full_name, (None, 0))
            display = nn_w or nn_c or full_name
            eff = produced / worked if worked > 0 else 0
            ws2.append([display, f"{worked:.2f}", f"{produced:.2f}", f"{eff:.2f}"])

        # --- Audit Log ---
        ws3 = wb.create_sheet("Audit Log")
        with get_connection() as conn:
            cur = conn.cursor()
            query3 = f"""
                SELECT c.date, r.ro_number, e.full_name, e.nickname, c.hours, c.note
                FROM credit_audit c
                JOIN employees e ON c.employee_id = e.id
                JOIN repair_orders r ON c.ro_id = r.id
                {where_c}
                ORDER BY c.id DESC
            """
            cur.execute(query3, params_c) if params_c else cur.execute(query3)
            rows = cur.fetchall()
        ws3.append(["Date", "RO#", "Employee", "Hours", "Note"])
        for date, ro_number, full_name, nickname, hours, note in rows:
            qdate = QDate.fromString(date.split()[0], "yyyy-MM-dd")
            date_fmt = qdate.toString("MM/dd/yyyy") if qdate.isValid() else date
            display = nickname or full_name
            ws3.append([date_fmt, ro_number, display, f"{hours:.2f}", note or ""])

        wb.save(file)
        QMessageBox.information(self, "Export", f"Reports exported to {file}")




